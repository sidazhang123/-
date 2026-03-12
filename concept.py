# -*- coding: utf-8 -*-
import logging
import json
import requests
import sqlite3
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import re

logger = logging.getLogger(__name__)    

EXCLUDE_BOARD_NAMES = ["融资融券","深股通","央国企改革","沪股通","预盈预增","创业板综","富时罗素","预亏预减","专精特新","标准普尔","机构重仓","MSCI中国","深成500","中证500","转债标的","微盘股","上证380","HS300_","QFII重仓","证金持股","参股银行","股权激励","养老金","上证180_","贬值受益","ST股","AH股","PPP模式","参股券商","参股新三板","深证100R","创业成份","百元股","参股保险","稀缺资源","昨日涨停_含一字","昨日涨停","基金重仓","AB股","IPO受益","央视50_","上证50_","参股期货","科创板做市股","银行","茅指数","股权转让","昨日触板","广电","宁组合","综合行业","彩票概念","社保重仓","昨日连板_含一字","租售同权","举牌","科创板做市商","退税商店","昨日连板","赛马概念"]



def get_all_stock_codes() -> list:
    """Get all stock codes from API and return as a list of stock_code strings."""
    logger.info("[get_all_stock_codes] Start fetching all stock codes from API")
    stock_codes = []
    session = requests.Session()
    BROWSER_HEADERS = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Connection': 'keep-alive',
    }
    session.headers.update(BROWSER_HEADERS)
    for num in range(1, 61):
        url = f"http://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQNodeData?num=200&sort=code&asc=0&node=hs_a&symbol=&_s_r_a=page&page={num}"
        try:
            response = session.get(url)
            data = json.loads(response.text)
        except Exception as e:
            logger.error(f"[get_all_stock_codes] Error fetching page {num}: {e}")
            continue
        if not data:
            continue
        for item in data:
            symbol = item.get('symbol', '')
            code = item.get('code', '')
            if symbol and code:
                stock_code = f"{code}.{symbol[:2]}".upper()
                stock_codes.append(stock_code)
    logger.info(f"[get_all_stock_codes] Fetched {len(stock_codes)} total stock codes")
    return stock_codes



def fetch_concept_data(stock_code):
    """
    Fetch concept data for a single stock_code from the Eastmoney API.
    Returns a list of tuples with the required fields, excluding those with BOARD_NAME in EXCLUDE_BOARD_NAMES.
    """
    url = f"https://datacenter.eastmoney.com/securities/api/data/get?type=RPT_F10_CORETHEME_BOARDTYPE&sty=ALL&filter=(SECUCODE%3D%22{stock_code}%22)&p=1&ps=&sr=1&st=BOARD_RANK&source=HSF10&client=PC"
    try:
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        results = data.get("result", {}).get("data", [])
        records = []
        for item in results:
            board_name = item.get("BOARD_NAME", "")
            if board_name in EXCLUDE_BOARD_NAMES:
                continue
            record = (
                item.get("SECURITY_CODE", ""),
                item.get("SECURITY_NAME_ABBR", ""),
                board_name,
                item.get("SELECTED_BOARD_REASON", "")
            )
            records.append(record)
        return records
    except Exception as e:
        logger.error(f"[fetch_concept_data] Error fetching {stock_code}: {e}")
        return []

def insert_concepts_to_db(records, db_path: str = 'concepts.db'):
    """
    Insert a list of concept records into the SQLite DB. Clears the table first.
    """
    logger.info(f"[insert_concepts_to_db] Inserting {len(records)} records into DB {db_path}")
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS concepts (
            SECURITY_CODE TEXT,
            SECURITY_NAME_ABBR TEXT,
            BOARD_NAME TEXT,
            SELECTED_BOARD_REASON TEXT
        )
    ''')
    cursor.execute('DELETE FROM concepts')
    if records:
        cursor.executemany(
            'INSERT INTO concepts (SECURITY_CODE, SECURITY_NAME_ABBR, BOARD_NAME, SELECTED_BOARD_REASON) VALUES (?, ?, ?, ?)',
            records
        )
        logger.info(f"[insert_concepts_to_db] Inserted {len(records)} records into concepts table.")
    else:
        logger.info("[insert_concepts_to_db] No records to insert into concepts table.")
    conn.commit()
    conn.close()
    logger.info(f"[insert_concepts_to_db] DB connection closed.")

def clean_excel_value(value):
    # Remove illegal characters for Excel XML
    if isinstance(value, str):
        # Remove control characters except for tab, newline, carriage return
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', value)
    return value

def export_to_xlsx(records):
    """
    Export records to an xlsx file with the specified header and formatting.
    """
    # Process third element of each item of records: add "(count of the element)" to it,
    # where count is the number of times the third element appears as a distinct one in records
    from collections import Counter

    third_elements = [rec[2] for rec in records]
    counts = Counter(third_elements)
    processed_records = []
    for rec in records:
        rec = list(rec)
        third_elem = rec[2]
        count = counts.get(third_elem, 0)
        rec[2] = f"{third_elem}({count})"
        processed_records.append(tuple(rec))
    processed_records.sort(key=lambda x: x[2])
    records = processed_records
    filename = datetime.now().strftime('%m%d') + '个股东财概念.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '概念板块'
    header = ['股票代码', '股票名称', '板块(含行业、概念，括号中为板块中个股数量)', '入选板块理由']
    ws.append(header)

    # Define a thin border style
    thin_side = Side(style='thin', color='000000')
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )
    # Set column widths 
    col_widths = [15, 15, 15, 120]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row bold
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.border = thin_border
    # Make header row filterable (add autofilter to header)
    ws.auto_filter.ref = ws.dimensions
    

    # Add data rows, cleaning each value
    for row in records:
        clean_row = [clean_excel_value(cell) for cell in row]
        ws.append(clean_row)

    # Center and wrap all cells and set border
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = align
            cell.border = thin_border

    wb.save(filename)
    logger.info(f"[export_to_xlsx] Exported {len(records)} records to {filename}")


def fetch_and_store_concepts(stock_codes: list, db_path: str = 'concepts.db'):
    """
    Use a thread pool to fetch concept data for each stock_code and store in SQLite DB.
    """
    logger.info(f"[fetch_and_store_concepts] Start fetching and storing concepts for {len(stock_codes)} stock codes.")
    all_records = []
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_concept_data, code) for code in stock_codes]
        for future in as_completed(futures):
            try:
                records = future.result()
                if records:
                    all_records.extend(records)
            except Exception as e:
                logger.error(f"[fetch_and_store_concepts] Error in future: {e}")
    logger.info(f"[fetch_and_store_concepts] Finished fetching. Total records to insert: {len(all_records)}")
    insert_concepts_to_db(all_records, db_path)
    logger.info(f"[fetch_and_store_concepts] All done.")
    # import pickle
    # with open('all_records.pkl', 'wb') as f:
    #     pickle.dump(all_records, f)
    logger.info(f"[fetch_and_store_concepts] Saved all_records to all_records.pkl")
    export_to_xlsx(all_records)

if __name__ == "__main__":
    # import os
    # import pickle

    # if os.path.exists('all_records.pkl'):
    #     logger.info("[main] all_records.pkl found. Loading and exporting to xlsx only.")
    #     with open('all_records.pkl', 'rb') as f:
    #         all_records = pickle.load(f)
    #     export_to_xlsx(all_records)
    #     logger.info("[main] Exported from cached all_records.pkl. Exiting.")
    #     exit(0)
    logging.basicConfig(level=logging.INFO)
    logger.info("[main] Program started.")
    stock_codes = get_all_stock_codes()
    logger.info(f"[main] Got {len(stock_codes)} stock codes. Starting concept fetch...")
    fetch_and_store_concepts(stock_codes)
    logger.info("[main] Program finished.")