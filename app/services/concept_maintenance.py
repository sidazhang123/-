"""
概念更新执行服务。

职责：
1. 从源库股票列表抓取东财概念与入选理由。
2. 对概念数据执行过滤、重试、并发抓取与全量刷新写库。
3. 输出可复用的任务摘要，供状态卡、日志和结果页查询使用。
"""

from __future__ import annotations

import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable
import re

import requests

from app.db.duckdb_utils import connect_duckdb_compatible
from app.services.maintenance_codecs import normalize_code_for_storage
from app.settings import (
    CONCEPT_EXCLUDE_BOARD_NAMES,
    CONCEPT_MAX_WORKERS,
    CONCEPT_REQUEST_TIMEOUT_SECONDS,
    CONCEPT_RETRY_COUNT,
)


class ConceptStopRequested(Exception):
    """
    概念任务收到停止请求时抛出的内部异常。
    """


class ConceptEmptyResultError(RuntimeError):
    """
    概念抓取结果为空且禁止覆盖旧表时抛出的异常。
    """


@dataclass(slots=True)
class ConceptRunSummary:
    """
    概念任务摘要。
    """

    steps_total: int
    steps_completed: int
    total_tasks: int
    success_tasks: int
    failed_tasks: int
    records_written: int
    filtered_records: int
    duration_seconds: float


class ConceptMaintenanceService:
    """
    概念更新执行器。
    """

    def __init__(
        self,
        *,
        source_db_path: Path | str,
        logger: Any,
        stop_checker: Callable[[], bool] | None = None,
        progress_reporter: Callable[[float, str | None, dict[str, object] | None], None] | None = None,
    ) -> None:
        """
        输入：
        1. source_db_path: 源库路径。
        2. logger: 结构化日志器。
        3. stop_checker: 停止检查函数。
        4. progress_reporter: 进度回调。
        输出：
        1. 无返回值。
        用途：
        1. 初始化概念任务运行依赖。
        边界条件：
        1. stop_checker/progress_reporter 为空时自动降级为 no-op。
        """

        self.source_db_path = Path(source_db_path)
        self.logger = logger
        self.stop_checker = stop_checker or (lambda: False)
        self.progress_reporter = progress_reporter or (lambda progress, phase=None, detail=None: None)
        self.timeout_seconds = float(CONCEPT_REQUEST_TIMEOUT_SECONDS)
        self.max_workers = max(1, int(CONCEPT_MAX_WORKERS))
        self.retry_count = max(0, int(CONCEPT_RETRY_COUNT))
        self.exclude_board_names = set(CONCEPT_EXCLUDE_BOARD_NAMES)
        self.http_session = self._build_http_session()

    @staticmethod
    def _build_http_session() -> requests.Session:
        """
        输入：
        1. 无。
        输出：
        1. 概念抓取专用 HTTP Session。
        用途：
        1. 为东财概念请求强制禁用环境代理，避免被系统 HTTP(S)_PROXY 污染。
        边界条件：
        1. 仅影响本服务内部请求，不修改全局环境变量。
        """

        session = requests.Session()
        session.trust_env = False
        return session

    @staticmethod
    def _clean_excel_value(value: object) -> object:
        if isinstance(value, str):
            return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", value)
        return value

    def _export_to_xlsx(
        self,
        records: list[tuple[str, str, str]],
        stock_name_map: dict[str, str],
    ) -> Path:
        """
        输入：
        1. records: 已去重的概念记录列表。
        2. stock_name_map: 股票代码到股票名称的映射。
        输出：
        1. 导出的 xlsx 文件路径。
        用途：
        1. 在概念整表更新成功后输出可人工复核的 Excel 文件。
        边界条件：
        1. 文件输出到源库同级目录；股票名称缺失时导出为空串。
        """

        from collections import Counter

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.utils import get_column_letter

        board_counts = Counter(board_name for _, board_name, _ in records)
        export_rows = []
        for code, board_name, reason in records:
            export_rows.append(
                (
                    str(code or "").strip(),
                    str(stock_name_map.get(code, "") or "").strip(),
                    f"{board_name}({board_counts.get(board_name, 0)})",
                    str(reason or "").strip(),
                )
            )
        export_rows.sort(key=lambda item: (item[2], item[0]))

        output_path = self.source_db_path.parent / (datetime.now().strftime("%m%d") + "个股东财概念.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "概念板块"
        ws.append(["股票代码", "股票名称", "板块(含行业、概念，括号中为板块中个股数量)", "入选板块理由"])

        thin_side = Side(style="thin", color="000000")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        for index, width in enumerate([15, 15, 24, 120], start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font
            cell.border = thin_border
        ws.auto_filter.ref = ws.dimensions

        for row in export_rows:
            ws.append([self._clean_excel_value(value) for value in row])

        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.alignment = align
                cell.border = thin_border

        wb.save(output_path)
        self._log_info("概念结果已导出到 Excel", {"xlsx_path": str(output_path), "records_exported": len(export_rows)})
        return output_path

    def _check_stop(self) -> None:
        """
        输入：
        1. 无。
        输出：
        1. 无返回值。
        用途：
        1. 在关键节点检查是否收到停止请求。
        边界条件：
        1. 收到停止请求时抛出内部异常终止流程。
        """

        if self.stop_checker():
            raise ConceptStopRequested("收到停止请求")

    def _log_info(self, message: str, detail: dict[str, object] | None = None) -> None:
        logger = self.logger
        if logger is None or not hasattr(logger, "info"):
            return
        logger.info(message, detail or {})

    def _log_error(self, message: str, detail: dict[str, object] | None = None) -> None:
        logger = self.logger
        if logger is None or not hasattr(logger, "error"):
            return
        logger.error(message, detail or {})

    def _connect_source_db(self):
        """
        输入：
        1. 无。
        输出：
        1. DuckDB 连接。
        用途：
        1. 打开源库并确保目录存在。
        边界条件：
        1. 连接异常由上层处理。
        """

        return connect_duckdb_compatible(self.source_db_path, read_only=False, retries=20, retry_sleep_seconds=0.05)

    def ensure_runtime_schema(self) -> None:
        """
        输入：
        1. 无。
        输出：
        1. 无返回值。
        用途：
        1. 确保源库中的股票概念表存在。
        边界条件：
        1. 仅创建缺失表，不清理历史数据。
        """

        with self._connect_source_db() as con:
            con.execute(
                """
                create table if not exists stock_concepts (
                    code varchar not null,
                    board_name varchar not null,
                    selected_reason varchar,
                    updated_at timestamp not null,
                    primary key (code, board_name)
                )
                """
            )

    def _load_stock_inputs(self) -> list[tuple[str, str]]:
        """
        输入：
        1. 无。
        输出：
        1. 股票代码与名称列表。
        用途：
        1. 从源库读取待抓取的股票池。
        边界条件：
        1. 无股票时返回空列表。
        """

        with self._connect_source_db() as con:
            rows = con.execute("select code, name from stocks order by code asc").fetchall()
        return [(str(code), str(name or "")) for code, name in rows if str(code or "").strip()]

    @staticmethod
    def _to_eastmoney_secu_code(storage_code: str) -> str:
        """
        输入：
        1. storage_code: 源库存储格式代码，如 sh.600000。
        输出：
        1. 东财接口所需代码，如 600000.SH。
        用途：
        1. 统一概念接口代码格式转换。
        边界条件：
        1. 非法输入返回空字符串。
        """

        token = str(storage_code or "").strip().lower()
        if "." not in token:
            return ""
        prefix, digits = token.split(".", 1)
        if not prefix or not digits:
            return ""
        return f"{digits}.{prefix.upper()}"

    def _fetch_concepts_for_stock(self, code: str, name: str) -> tuple[list[tuple[str, str, str]], int]:
        """
        输入：
        1. code: 源库格式股票代码。
        2. name: 股票名称。
        输出：
        1. 过滤后的概念记录列表和过滤掉的条数。
        用途：
        1. 抓取单只股票的概念明细。
        边界条件：
        1. 请求失败时抛出异常，由上层决定是否重试。
        """

        secu_code = self._to_eastmoney_secu_code(code)
        if not secu_code:
            raise ValueError(f"非法股票代码: {code}")
        url = (
            "https://datacenter.eastmoney.com/securities/api/data/get"
            f"?type=RPT_F10_CORETHEME_BOARDTYPE&sty=ALL&filter=(SECUCODE%3D%22{secu_code}%22)"
            "&p=1&ps=&sr=1&st=BOARD_RANK&source=HSF10&client=PC"
        )

        last_exc: Exception | None = None
        for attempt in range(self.retry_count + 1):
            self._check_stop()
            try:
                response = self.http_session.get(url, timeout=self.timeout_seconds)
                response.raise_for_status()
                payload = response.json()
                results = payload.get("result", {}).get("data", []) or []
                records: list[tuple[str, str, str]] = []
                filtered = 0
                for item in results:
                    board_name = str(item.get("BOARD_NAME") or "").strip()
                    if not board_name:
                        continue
                    if board_name in self.exclude_board_names:
                        filtered += 1
                        continue
                    reason = str(item.get("SELECTED_BOARD_REASON") or "").strip()
                    records.append((normalize_code_for_storage(code), board_name, reason))
                return records, filtered
            except Exception as exc:  # pragma: no cover - 重试分支依赖网络条件
                last_exc = exc
                if attempt >= self.retry_count:
                    raise
                time.sleep(min(0.8, 0.2 * (attempt + 1)))
        raise RuntimeError(str(last_exc) if last_exc else "未知概念抓取错误")

    @staticmethod
    def _merge_reason_text(existing: str, incoming: str) -> str:
        """
        输入：
        1. existing: 已保留的入选理由。
        2. incoming: 新进入的入选理由。
        输出：
        1. 合并后的入选理由文本。
        用途：
        1. 在同股同概念重复出现时保留去重后的理由信息。
        边界条件：
        1. 任一侧为空时返回另一侧；重复文本只保留一次。
        """

        existing_text = str(existing or "").strip()
        incoming_text = str(incoming or "").strip()
        if not existing_text:
            return incoming_text
        if not incoming_text or incoming_text == existing_text:
            return existing_text

        parts: list[str] = []
        seen: set[str] = set()
        for token in [existing_text, *incoming_text.split("；")]:
            normalized = str(token or "").strip()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            parts.append(normalized)
        return "；".join(parts)

    def _deduplicate_records(self, records: list[tuple[str, str, str]]) -> list[tuple[str, str, str]]:
        """
        输入：
        1. records: 原始概念记录列表。
        输出：
        1. 以 (code, board_name) 去重后的记录列表。
        用途：
        1. 屏蔽上游接口重复返回的同股同概念数据，避免整表刷新时主键冲突。
        边界条件：
        1. 保留首次出现顺序；重复项会合并 selected_reason。
        """

        merged: dict[tuple[str, str], str] = {}
        ordered_keys: list[tuple[str, str]] = []
        for code, board_name, reason in records:
            key = (str(code or "").strip(), str(board_name or "").strip())
            if not key[0] or not key[1]:
                continue
            if key not in merged:
                merged[key] = str(reason or "").strip()
                ordered_keys.append(key)
                continue
            merged[key] = self._merge_reason_text(merged[key], str(reason or "").strip())
        return [(code, board_name, merged[(code, board_name)]) for code, board_name in ordered_keys]

    def _replace_all_records(self, records: list[tuple[str, str, str]]) -> int:
        """
        输入：
        1. records: 概念记录列表。
        输出：
        1. 实际写入条数。
        用途：
        1. 以全量覆盖方式刷新 stock_concepts。
        边界条件：
        1. 调用方应先确保非空；本函数仅负责全量覆盖已准备好的数据。
        """

        deduplicated_records = self._deduplicate_records(records)
        dropped_duplicates = max(0, len(records) - len(deduplicated_records))
        if dropped_duplicates > 0:
            self._log_info(
                "概念写库前已去重重复记录",
                {
                    "input_records": len(records),
                    "deduplicated_records": len(deduplicated_records),
                    "dropped_duplicates": dropped_duplicates,
                },
            )

        now = datetime.now()
        with self._connect_source_db() as con:
            con.execute("delete from stock_concepts")
            if deduplicated_records:
                con.executemany(
                    """
                    insert into stock_concepts(code, board_name, selected_reason, updated_at)
                    values (?, ?, ?, ?)
                    """,
                    [(code, board_name, reason, now) for code, board_name, reason in deduplicated_records],
                )
        return len(deduplicated_records)

    def run_update(self) -> ConceptRunSummary:
        """
        输入：
        1. 无。
        输出：
        1. 概念更新摘要。
        用途：
        1. 执行一次完整概念刷新。
        边界条件：
        1. 运行中若收到停止请求，会抛出 ConceptStopRequested。
        """

        started_at = time.monotonic()
        self.ensure_runtime_schema()
        self.progress_reporter(1.0, "prepare", {"stage": "load_stocks"})
        stock_inputs = self._load_stock_inputs()
        stock_name_map = {str(code or "").strip(): str(name or "").strip() for code, name in stock_inputs}
        total_tasks = len(stock_inputs)
        self._log_info("概念更新任务准备完成", {"total_stocks": total_tasks, "max_workers": self.max_workers})

        if total_tasks <= 0:
            raise ConceptEmptyResultError(
                "概念更新结果为空：源库无可抓取股票，已保留旧概念数据"
            )

        collected_records: list[tuple[str, str, str]] = []
        filtered_records = 0
        success_tasks = 0
        failed_tasks = 0
        self.progress_reporter(3.0, "running", {"stage": "fetch_concepts", "total_tasks": total_tasks})

        with ThreadPoolExecutor(max_workers=self.max_workers, thread_name_prefix="concept-fetch") as executor:
            futures = {
                executor.submit(self._fetch_concepts_for_stock, code, name): (code, name)
                for code, name in stock_inputs
            }
            for index, future in enumerate(as_completed(futures), start=1):
                self._check_stop()
                code, name = futures[future]
                try:
                    records, filtered = future.result()
                    success_tasks += 1
                    filtered_records += int(filtered)
                    if records:
                        collected_records.extend(records)
                except Exception as exc:
                    failed_tasks += 1
                    self._log_error(
                        "抓取股票概念失败",
                        {"code": code, "name": name, "error": f"{type(exc).__name__}: {exc}"},
                    )
                progress = 3.0 + (index * 92.0 / max(1, total_tasks))
                self.progress_reporter(
                    progress,
                    "running",
                    {
                        "stage": "fetch_concepts",
                        "processed_tasks": index,
                        "total_tasks": total_tasks,
                        "records_collected": len(collected_records),
                    },
                )

        self._check_stop()
        if not collected_records:
            raise ConceptEmptyResultError(
                "概念更新结果为空：本次未抓到任何概念记录，已保留旧概念数据"
            )
        deduplicated_records = self._deduplicate_records(collected_records)
        self.progress_reporter(97.0, "writing", {"stage": "replace_table", "records": len(collected_records)})
        records_written = self._replace_all_records(deduplicated_records)
        self._export_to_xlsx(deduplicated_records, stock_name_map)
        summary = ConceptRunSummary(
            steps_total=2,
            steps_completed=2,
            total_tasks=total_tasks,
            success_tasks=success_tasks,
            failed_tasks=failed_tasks,
            records_written=records_written,
            filtered_records=filtered_records,
            duration_seconds=round(time.monotonic() - started_at, 3),
        )
        self._log_info(
            "概念更新任务完成",
            {
                "total_tasks": total_tasks,
                "success_tasks": success_tasks,
                "failed_tasks": failed_tasks,
                "records_written": records_written,
                "filtered_records": filtered_records,
                "duration_seconds": summary.duration_seconds,
            },
        )
        self.progress_reporter(100.0, "done", {"stage": "completed", "records_written": records_written})
        return summary